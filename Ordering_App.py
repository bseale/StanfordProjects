import pyperclip

from openpyxl import load_workbook

wb = load_workbook('A. Lab Order Request.xlsx')

ws = wb.active

run_again = ''
while run_again == '':
    x = input('Line #: ')
    lm = ws["C" + str(x)].value
    item = ws["K" + str(x)].value
    purpose = ws["P" + str(x)].value
    funding_code = ws["I" + str(x)].value
    order_status = ws["G" + str(x)].value
    part_number = str(ws["L" + str(x)].value)
    quantity_order = str(ws["M" + str(x)].value)
    price_order = str(ws["N" + str(x)].value)

    if lm in "BS":
        lm_full = "Brandon Seale"
    elif lm in "JM":
        lm_full = "Jens Magnusson"
    elif lm in "MH":
        lm_full = "Mengting Han"
    elif lm in "XX":
        lm_full = "Xiaoshu Xu"
    elif lm in "LZ":
        lm_full = "Leiping Zeng"
    elif lm in "YZ":
        lm_full = "Yanyu Zhu"
    elif lm in "VT":
        lm_full = "Victor Tieu"
    elif lm in "CC":
        lm_full = "Crystal Chen"
    elif lm in "XC":
        lm_full = "Xinyi Chen"
    elif lm in "SC":
        lm_full = "Sa Cai"
    elif lm in "GR":
        lm_full = "Goldie Roth"
    elif lm in "MCC":
        lm_full = "Marvin Collins"
    elif lm in "YS":
        lm_full = "Yinglin Situ"
    elif lm in "LM":
        lm_full = "Leanne Miles"
    else:
        lm_full = "xxxxxxxx"

    print(lm_full)

    bizpurpose = f'On behalf of the Qi Lab, {lm_full} ordered {item} for: ({purpose}).Used in research funded by {funding_code}'
    print(bizpurpose)
    pyperclip.copy(bizpurpose)

    print(f'Part Number: {part_number}')
    print(f'Quantity: {quantity_order}')
    print(f'Price: {price_order}')

    run_again = input("Press Enter to Run Code Again").upper()
